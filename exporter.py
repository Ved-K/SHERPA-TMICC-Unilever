from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import store
from risk_matrix import lookup_rating

BACKUP_HEADERS = [
    "", "", "", "", "", "", "", "",  # spacing
    "Line Code", "Machine Code", "Task Code", "Task ID", "Step ID",
    "Operation Category", "Phases", "Probability", "Severity",
    "Eng Controls", "Admin Controls", "Hazard Text",
]

EXPORT_HEADERS = [
    "Area", "Job/Activity", "Task/Step", "Hazard Description",
    "Existing Controls/Defences", "Probability (P)", "Severity (S)", "Additional Info",
] + BACKUP_HEADERS


CATEGORY_MAP = {
    "Normal Operations": "NO",
    "Abnormal Operations": "AO",
    "Emergency": "EM",
}

STATUS_MAP = {
    "Startup": "ST",
    "Shutdown": "SD",
    "Running": "RN",
}


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(60, max(12, max_len + 2))


def export_to_excel(con, out_path: str, rating_map: Dict[Any, Any]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "ProgramOutputData"

    EXPORT_HEADERS = [
        "Area", "Job/Activity", "Task/Step", "Hazard Description",
        "Existing Controls/Defences", "Probability (P)", "Severity (S)", "Additional Info",
        "", "", "", "", "", "", "", "",  # spacing
        "Line Code", "Line Name", "Machine Code", "Machine Name",
        "Task Code", "Task ID", "Task Name",
        "Step ID", "Step No", "Step Desc",
        "Operation Category", "Phases",
        "Eng Controls", "Admin Controls", "Hazard Text",
        "Probability Code", "Severity Code", "Risk Rating"
    ]
    ws.append(EXPORT_HEADERS)

    line_map, machine_map = {}, {}

    lines = store.list_lines(con)
    for li, ln in enumerate(lines, start=1):
        line_code = f"L{li}"
        line_map[line_code] = ln["name"]

        machines = store.list_machines(con, ln["id"])
        for mi, m in enumerate(machines, start=1):
            mach_code = f"M{mi}"
            machine_map[mach_code] = m["name"]

            tasks = store.list_tasks(con, m["id"])
            for ti, t in enumerate(tasks, start=1):
                cat = CATEGORY_MAP.get(t["operation_category"], t["operation_category"][:2].upper())
                status = next((abbr for name, abbr in STATUS_MAP.items() if name.lower() in t["name"].lower()), "RN")
                task_code = f"{line_code}-{mach_code}-T{ti}-{cat}-{status}"

                # Task header row (blank for step data)
                ws.append([
                    ln["name"], task_code, t["name"], "", "", "", "", "",
                    "", "", "", "", "", "", "", "",  # spacing
                    line_code, ln["name"], mach_code, m["name"],
                    task_code, t["id"], t["name"],
                    "", "", "", t["operation_category"], ", ".join(t["phases"]),
                    "", "", "", "", "", ""
                ])

                steps = store.list_steps(con, t["id"])
                for si, s in enumerate(steps, start=1):
                    job_name = f"{line_code}-{mach_code}-T{ti}-S{si}-{cat}-{status}"
                    combined_controls = "; ".join(filter(None, [s["eng_controls"], s["admin_controls"]]))
                    rating = lookup_rating(float(s["probability_code"]), float(s["severity_code"]), rating_map)
                    ws.append([
                        ln["name"], job_name, s["step_desc"], s["hazard_text"],
                        combined_controls, s["probability_code"], s["severity_code"], "",
                        "", "", "", "", "", "", "", "",
                        line_code, ln["name"], mach_code, m["name"],
                        task_code, t["id"], t["name"],
                        s["id"], s["step_no"], s["step_desc"],
                        t["operation_category"], ", ".join(t["phases"]),
                        s["eng_controls"], s["admin_controls"], s["hazard_text"],
                        s["probability_code"], s["severity_code"], rating
                    ])
                ws.append([])  # space between tasks
            ws.append([])  # space between machines
        ws.append([])  # space between lines

    # --- Styling ---
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.row == 1:
                cell.font = Font(bold=True)

    _autosize(ws)
    wb.save(out_path)

    # --- Legend PDF ---
    legend_path = Path(out_path).with_name("SHERPA_Legend.pdf")
    _export_legend_pdf(legend_path, line_map, machine_map)
    return str(out_path)


def _export_legend_pdf(out_path: Path, line_map: Dict[str, str], machine_map: Dict[str, str]):
    from reportlab.lib.units import cm

    c = canvas.Canvas(str(out_path), pagesize=A4)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2 * cm, 27 * cm, "SHERPA Legend")

    c.setFont("Helvetica", 11)
    y = 25.5 * cm
    c.drawString(2 * cm, y, "Line Codes:")
    y -= 0.5 * cm
    for k, v in line_map.items():
        c.drawString(2.5 * cm, y, f"{k} — {v}")
        y -= 0.4 * cm

    y -= 0.5 * cm
    c.drawString(2 * cm, y, "Machine Codes:")
    y -= 0.5 * cm
    for k, v in machine_map.items():
        c.drawString(2.5 * cm, y, f"{k} — {v}")
        y -= 0.4 * cm

    y -= 0.5 * cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2 * cm, y, "Category & Status Codes:")
    y -= 0.5 * cm
    c.setFont("Helvetica", 11)
    for cat, abbr in CATEGORY_MAP.items():
        c.drawString(2.5 * cm, y, f"{abbr} — {cat}")
        y -= 0.4 * cm
    for stat, abbr in STATUS_MAP.items():
        c.drawString(2.5 * cm, y, f"{abbr} — {stat}")
        y -= 0.4 * cm

    c.save()
