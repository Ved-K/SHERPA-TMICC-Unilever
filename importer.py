from openpyxl import load_workbook
from io import BytesIO
import json
import store

def _cell(row, idx, default=""):
    # safe getter for variable-length tuples
    if idx < 0 or idx >= len(row):
        return default
    v = row[idx]
    return default if v is None else v

def import_from_excel(con, file_obj) -> None:
    wb = load_workbook(file_obj, data_only=True)
    ws = wb["ProgramOutputData"]

    # ---- Find backup columns by header name (robust: no hardcoded indexes) ----
    header = [c if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {str(name).strip(): i for i, name in enumerate(header)}

    def c(name: str) -> int:
        # raises early if your export changes so you notice instantly
        if name not in col:
            raise KeyError(f"Missing column in Excel: '{name}'")
        return col[name]

    # These must exist in your backup section (adjust names if yours differ)
    I_LINE_CODE   = c("Line Code")
    I_LINE_NAME   = c("Line Name")
    I_MACH_CODE   = c("Machine Code")
    I_MACH_NAME   = c("Machine Name")
    I_TASK_CODE   = c("Task Code")
    I_TASK_NAME   = c("Task Name")
    I_OP_CAT      = c("Operation Category")
    I_PHASES      = c("Phases")
    I_STEP_NO     = c("Step No")
    I_STEP_DESC   = c("Step Desc")
    I_HAZARDS     = c("Hazard Text")
    I_ENG         = c("Eng Controls")
    I_ADMIN       = c("Admin Controls")
    I_P           = c("Probability Code")
    I_S           = c("Severity Code")

    with con:  # one transaction
        for row in ws.iter_rows(min_row=2, values_only=True):
            line_code = str(_cell(row, I_LINE_CODE, "")).strip()
            line_name = str(_cell(row, I_LINE_NAME, "")).strip()

            mach_code = str(_cell(row, I_MACH_CODE, "")).strip()
            mach_name = str(_cell(row, I_MACH_NAME, "")).strip()

            task_code = str(_cell(row, I_TASK_CODE, "")).strip()
            task_name = str(_cell(row, I_TASK_NAME, "")).strip()
            op_cat    = str(_cell(row, I_OP_CAT, "")).strip()

            phases_raw = _cell(row, I_PHASES, "")
            phases = [p.strip() for p in str(phases_raw).split(",") if p.strip()]

            step_no_raw = _cell(row, I_STEP_NO, None)
            step_desc   = str(_cell(row, I_STEP_DESC, "")).strip()

            hazards     = str(_cell(row, I_HAZARDS, "")).strip()
            eng_ctrl    = str(_cell(row, I_ENG, "")).strip()
            admin_ctrl  = str(_cell(row, I_ADMIN, "")).strip()

            p = _cell(row, I_P, None)
            s = _cell(row, I_S, None)

            # skip empty dump rows / separators
            if not any([line_name, mach_name, task_name, step_desc, hazards, eng_ctrl, admin_ctrl]):
                continue

            # ---------------- LINE ----------------
            # Use backup code + name (both). If code blank, generate code.
            lid = store.ensure_line(con, line_code, line_name)

            # ---------------- MACHINE ----------------
            mid = store.ensure_machine(con, lid, mach_code, mach_name)

            # ---------------- TASK ----------------
            tid = store.ensure_task(con, mid, task_code, task_name, op_cat, phases)

            # ---------------- STEP ----------------
            # Only insert step if it actually has a step number or description
            if step_no_raw not in (None, "", 0) or step_desc:
                try:
                    step_no = int(step_no_raw or 0)
                except Exception:
                    step_no = 0

                if step_no <= 0:
                    # last resort fallback: append to end for that task
                    step_no = store.next_step_no(con, tid)

                p_clean = str(p if p not in (None, "") else "0")
                s_clean = str(s if s not in (None, "") else "0")

                # ensure_step should accept eng+admin separately (recommended)
                store.ensure_step(
                    con,
                    task_id=tid,
                    step_no=step_no,
                    desc=step_desc,
                    hazard=hazards,
                    eng_controls=eng_ctrl,
                    admin_controls=admin_ctrl,
                    p=p_clean,
                    s=s_clean,
                )
