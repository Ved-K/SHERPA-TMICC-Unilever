from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, List, Tuple
import openpyxl


@dataclass(frozen=True)
class Option:
    label: str
    code: float


def load_risk_matrix(xlsx_path: str) -> Tuple[List[Option], List[Option], Dict[str, str]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["Risk Matrix"]

    sev_cols = ["D", "E", "F", "G", "H", "I"]

    severity: List[Option] = []
    for col in sev_cols:
        label = ws[f"{col}5"].value
        code = ws[f"{col}6"].value
        if label is None or code is None:
            continue
        severity.append(Option(label=str(label).strip(), code=float(code)))

    probability: List[Option] = []
    for r in range(7, 13):
        label = ws[f"B{r}"].value
        code = ws[f"C{r}"].value
        if label is None or code is None:
            continue
        probability.append(Option(label=str(label).strip(), code=float(code)))

    rating_map: Dict[str, str] = {}
    for r in range(7, 13):
        p_code = ws[f"C{r}"].value
        if p_code is None:
            continue
        for col in sev_cols:
            s_code = ws[f"{col}6"].value
            rating = ws[f"{col}{r}"].value
            if s_code is None or rating is None:
                continue
            rating_map[f"P{float(p_code)}_S{float(s_code)}"] = str(rating).strip()

    return probability, severity, rating_map


def lookup_rating(prob_code: float, sev_code: float, rating_map: Dict[str, str]) -> str:
    return rating_map.get(f"P{float(prob_code)}_S{float(sev_code)}", "Unrated")
